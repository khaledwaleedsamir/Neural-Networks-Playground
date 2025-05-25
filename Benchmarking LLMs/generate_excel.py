from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "Q&A Benchmark"

# Define column headers and widths
headers = [
    "Arabic Dialect", "Question Topic", "Question", "Gold Answer", "LLM Answer",
    "BLEU", "ROUGE", "Accuracy (1–5)", "Fluency", "Relevance", "Hallucination?"
]
column_widths = [18, 20, 40, 40, 40, 10, 10, 18, 10, 12, 14]

# Merge headers for evaluation categories
ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)   # Automatic Evaluation
ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=11)  # Human Evaluation

# Set titles for merged cells
ws.cell(row=1, column=6, value="Automatic Evaluation").alignment = Alignment(horizontal='center', vertical='center')
ws.cell(row=1, column=8, value="Human Evaluation").alignment = Alignment(horizontal='center', vertical='center')

# Apply header row (row 2)
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_num, value=header)

# Apply formatting to header rows
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for row in [1, 2]:
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Adjust column widths
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Save the file
file_path = "arabic_qa_template.xlsx"
wb.save(file_path)
print(f"✅ New Excel file created: {file_path}")
